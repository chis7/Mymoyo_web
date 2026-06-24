from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError


def clean_text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def clean_coordinate(value, max_absolute_value):
    value = clean_text(value)
    if value is None:
        return None
    try:
        coordinate = Decimal(value)
    except InvalidOperation:
        return None
    if abs(coordinate) > max_absolute_value:
        return None
    return coordinate


def clean_facility_type(value):
    value = clean_text(value)
    if value is None:
        return None
    normalized = value.lower()
    if normalized in {'hub', 'spoke'}:
        return normalized
    if normalized in {'n/a', 'na', 'not applicable'}:
        return 'na'
    return None


class Command(BaseCommand):
    help = 'Import provinces, districts and facilities from an Excel file (xlsx).'

    def add_arguments(self, parser):
        parser.add_argument('file', help='Path to the Excel file (.xlsx)')

    def handle(self, *args, **options):
        path = options['file']
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl is required: pip install openpyxl')

        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f'Failed to open Excel file: {e}')

        sheet = wb.active
        headers = []
        rows = sheet.iter_rows(values_only=True)
        try:
            first = next(rows)
        except StopIteration:
            raise CommandError('Excel file is empty')

        headers = [str(h).strip().lower() if h is not None else '' for h in first]

        # detect columns
        col_map = {}
        for idx, h in enumerate(headers):
            if 'province' in h:
                col_map['province'] = idx
            if 'district' in h:
                col_map['district'] = idx
            if 'site type' in h or h in {'hub/spoke', 'hub spoke', 'facility type'}:
                col_map.setdefault('facility_type', idx)
            if h == 'hub':
                col_map.setdefault('hub', idx)
            # common facility/name headers
            if (
                'code' not in h and
                (
                    h in {'facility', 'health facility', 'facility name', 'site name', 'name'} or
                    'health facility' in h
                )
            ):
                col_map.setdefault('facility', idx)
            # possible code columns
            if 'facility code' in h or 'mfl' in h or 'hims' in h:
                col_map['code'] = idx
            elif 'code' in h and 'hub code' not in h:
                col_map.setdefault('code', idx)
            if 'level' in h or 'facility level' in h or h == 'type':
                col_map.setdefault('level', idx)
            if 'latitude' in h or h == 'lat':
                col_map.setdefault('latitude', idx)
            if 'longitude' in h or h in {'lng', 'lon', 'long'}:
                col_map.setdefault('longitude', idx)

        if not {'province', 'district', 'facility'}.issubset(set(col_map.keys())):
            raise CommandError('Could not detect required columns (province, district, facility) in the sheet header')

        from locations.models import Province, District, Facility
        created = {'provinces': 0, 'districts': 0, 'facilities': 0}
        updated_facilities = 0
        hub_spoke_links = []

        for row in rows:
            province_name = row[col_map['province']]
            district_name = row[col_map['district']]
            facility_name = row[col_map['facility']]
            if not (province_name and district_name and facility_name):
                continue
            province_name = str(province_name).strip()
            district_name = str(district_name).strip()
            facility_name = str(facility_name).strip()

            province_obj, p_created = Province.objects.get_or_create(name=province_name)
            if p_created:
                created['provinces'] += 1

            district_obj, d_created = District.objects.get_or_create(name=district_name, province=province_obj)
            if d_created:
                created['districts'] += 1

            defaults = {}
            if 'code' in col_map:
                defaults['code'] = clean_text(row[col_map['code']])
            if 'level' in col_map:
                defaults['level'] = clean_text(row[col_map['level']])
            if 'facility_type' in col_map:
                facility_type = clean_facility_type(row[col_map['facility_type']])
                if facility_type is not None:
                    defaults['facility_type'] = facility_type
            if 'latitude' in col_map:
                defaults['latitude'] = clean_coordinate(row[col_map['latitude']], 90)
            if 'longitude' in col_map:
                defaults['longitude'] = clean_coordinate(row[col_map['longitude']], 180)
            facility_obj, f_created = Facility.objects.update_or_create(
                name=facility_name,
                district=district_obj,
                defaults=defaults,
            )
            if defaults.get('facility_type') == Facility.FACILITY_TYPE_SPOKE and 'hub' in col_map:
                hub_name = clean_text(row[col_map['hub']])
                if hub_name:
                    hub_spoke_links.append((province_name, district_name, hub_name, facility_obj.pk))
            if f_created:
                created['facilities'] += 1
            else:
                updated_facilities += 1

        for province_name, district_name, hub_name, spoke_pk in hub_spoke_links:
            hub = Facility.objects.filter(
                name__iexact=hub_name,
                district__name__iexact=district_name,
                district__province__name__iexact=province_name,
            ).first()
            if hub:
                hub.facility_type = Facility.FACILITY_TYPE_HUB
                hub.save(update_fields=['facility_type'])
                Facility.objects.filter(pk=spoke_pk).update(hub=hub)

        wb.close()

        self.stdout.write(self.style.SUCCESS(
            f"Import complete - provinces: {created['provinces']}, districts: {created['districts']}, "
            f"facilities: {created['facilities']}, updated facilities: {updated_facilities}"
        ))
