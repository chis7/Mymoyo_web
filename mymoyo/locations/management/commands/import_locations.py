from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError


def clean_text(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def clean_coordinate(value, max_absolute_value):
    value = clean_text(value)
    if value is None:
        return None
    try:
        coordinate = Decimal(value)
    except InvalidOperation:
        return None
    if abs(coordinate) > max_absolute_value:
        return None
    return coordinate


class Command(BaseCommand):
    help = 'Import provinces, districts and facilities from an Excel file (xlsx).'

    def add_arguments(self, parser):
        parser.add_argument('file', help='Path to the Excel file (.xlsx)')

    def handle(self, *args, **options):
        path = options['file']
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl is required: pip install openpyxl')

        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f'Failed to open Excel file: {e}')

        sheet = wb.active
        headers = []
        rows = sheet.iter_rows(values_only=True)
        try:
            first = next(rows)
        except StopIteration:
            raise CommandError('Excel file is empty')

        headers = [str(h).strip().lower() if h is not None else '' for h in first]

        # detect columns
        col_map = {}
        for idx, h in enumerate(headers):
            if 'province' in h:
                col_map['province'] = idx
            if 'district' in h:
                col_map['district'] = idx
            # common facility/name headers
            if 'facility' in h or 'health facility' in h or 'facility name' in h or h == 'name' or 'site name' in h:
                col_map['facility'] = idx
            # possible code columns
            if 'mfl' in h or 'code' in h or 'facility code' in h or 'hims' in h:
                col_map.setdefault('code', idx)
            if 'level' in h or 'facility level' in h or 'type' in h:
                col_map.setdefault('level', idx)
            if 'latitude' in h or h == 'lat':
                col_map.setdefault('latitude', idx)
            if 'longitude' in h or h in {'lng', 'lon', 'long'}:
                col_map.setdefault('longitude', idx)

        if not {'province', 'district', 'facility'}.issubset(set(col_map.keys())):
            raise CommandError('Could not detect required columns (province, district, facility) in the sheet header')

        from locations.models import Province, District, Facility
        created = {'provinces': 0, 'districts': 0, 'facilities': 0}
        updated_facilities = 0

        for row in rows:
            province_name = row[col_map['province']]
            district_name = row[col_map['district']]
            facility_name = row[col_map['facility']]
            if not (province_name and district_name and facility_name):
                continue
            province_name = str(province_name).strip()
            district_name = str(district_name).strip()
            facility_name = str(facility_name).strip()

            province_obj, p_created = Province.objects.get_or_create(name=province_name)
            if p_created:
                created['provinces'] += 1

            district_obj, d_created = District.objects.get_or_create(name=district_name, province=province_obj)
            if d_created:
                created['districts'] += 1

            defaults = {}
            if 'code' in col_map:
                defaults['code'] = clean_text(row[col_map['code']])
            if 'level' in col_map:
                defaults['level'] = clean_text(row[col_map['level']])
            if 'latitude' in col_map:
                defaults['latitude'] = clean_coordinate(row[col_map['latitude']], 90)
            if 'longitude' in col_map:
                defaults['longitude'] = clean_coordinate(row[col_map['longitude']], 180)
            facility_obj, f_created = Facility.objects.update_or_create(
                name=facility_name,
                district=district_obj,
                defaults=defaults,
            )
            if f_created:
                created['facilities'] += 1
            else:
                updated_facilities += 1

        self.stdout.write(self.style.SUCCESS(
            f"Import complete - provinces: {created['provinces']}, districts: {created['districts']}, "
            f"facilities: {created['facilities']}, updated facilities: {updated_facilities}"
        ))
